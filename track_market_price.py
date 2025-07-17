import os
import time

import pandas as pd
from datetime import datetime  # 인증 토큰 발급 함수

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font

from utils.stock_utils import fetch_daily_chart, get_token
from utils.telegram import send_to_telegram

load_dotenv()

# 봇 토큰 설정
BOT_TOKEN = os.getenv('TOKEN')
GROUP_ID = os.getenv('GROUP_ID')

TODAY = datetime.today().strftime("%Y%m%d")

def get_today_closes(token: str, codes: list[str]) -> dict[str, int]:
    closes = {}

    for raw_code in codes:
        code = f"{int(raw_code):06d}"  # 6자리 문자열 보정
        print(f"📌 요청: {code}")

        df = fetch_daily_chart(token, code, base_date=TODAY)
        if df.empty:
            print(f"⚠️ {code}: 데이터 없음")
            continue

        try:
            close_price = int(float(df.iloc[0]["cur_prc"] or 0))
            closes[code] = close_price
            print(f"✅ {code}: {close_price}")
            time.sleep(0.2)
        except Exception as e:
            print(f"⚠️ {code} 종가 추출 실패: {e}")

    return closes


def update_history_file(base_file_path: str, closes: dict[str, int]):
    today_col = datetime.today().strftime("%Y-%m-%d")

    # 기존 파일 불러오기
    df_history = pd.read_excel(base_file_path, dtype={"종목코드": str})

    if "종목코드" not in df_history.columns:
        raise ValueError("❌ '종목코드' 컬럼이 필요합니다.")

    # 종가 컬럼이 없다면 추가
    if today_col not in df_history.columns:
        df_history[today_col] = None

    # ✅ 종가 업데이트: 종목코드 기준으로만 넣음 (인덱스 사용 안함)
    for idx, row in df_history.iterrows():
        raw_code = row.get("종목코드")

        try:
            code = f"{int(str(raw_code).strip().split('.')[0]):06d}"  # 문자열화 → 공백제거 → 소수점 제거 → 6자리
        except Exception as e:
            print(f"⚠️ 종목코드 변환 실패 (row {idx}): {raw_code} -> {e}")
            continue

        if code in closes:
            df_history.at[idx, today_col] = closes[code]
            print(f"✅ {code}: 종가 {closes[code]} 삽입 완료 (row {idx})")
        else:
            print(f"⚠️ {code} not in closes")

    df_history.to_excel(base_file_path, index=False)
    print(f"✅ 저장 완료: {base_file_path}")

    # ✅ 색상 적용
    wb = load_workbook(base_file_path)
    ws = wb.active

    # 열 인덱스 계산 (A=1부터 시작)
    col_index_map = {cell.value: cell.column for cell in ws[1]}
    if today_col not in col_index_map:
        print("⚠️ 오늘 컬럼을 찾을 수 없습니다.")
        return

    today_col_idx = col_index_map[today_col]
    all_dates = [col for col in df_history.columns if col not in ["종목코드", "종목명"]]
    all_dates_sorted = sorted(all_dates)

    if len(all_dates_sorted) < 2:
        print("ℹ️ 비교할 어제 데이터가 없어 색상 적용 생략")
        return

    yesterday_col = all_dates_sorted[-2]
    yesterday_col_idx = col_index_map[yesterday_col]

    for row_idx in range(2, ws.max_row + 1):  # 헤더 제외
        try:
            today_val = float(ws.cell(row=row_idx, column=today_col_idx).value)
            yesterday_val = float(ws.cell(row=row_idx, column=yesterday_col_idx).value)

            if today_val > yesterday_val:
                ws.cell(row=row_idx, column=today_col_idx).font = Font(color="FF0000")  # 빨강
            elif today_val < yesterday_val:
                ws.cell(row=row_idx, column=today_col_idx).font = Font(color="0000FF")  # 파랑
        except (TypeError, ValueError):
            continue  # 값이 없거나 잘못된 경우 생략

    wb.save(base_file_path)
    print("🎨 색상 적용 완료")


def update_excel_with_prices(file_path: str, closes: dict[str, int]):
    today_col = datetime.today().strftime("%Y-%m-%d")

    wb = load_workbook(file_path)
    ws = wb.active

    # ✅ 헤더에 오늘 날짜 열 추가
    print("max_col", ws.max_column)
    max_col = ws.max_column + 1
    ws.cell(row=1, column=max_col, value=today_col)

    # ✅ 종가 삽입 (종목코드 열은 1열이라고 가정)
    for row in range(2, ws.max_row + 1):  # ✅ 반드시 마지막 행 포함 → range(... + 1)
        code_cell = ws.cell(row=row, column=1)
        raw_code = code_cell.value

        try:
            # ✅ 견고한 종목코드 파싱: strip + 소수점 제거 + 6자리 보정
            code = f"{int(str(raw_code).strip().split('.')[0]):06d}"
        except Exception as e:
            print(f"⚠️ row {row} 종목코드 파싱 실패: {raw_code} -> {e}")
            continue

        if code in closes:
            price = closes[code]
            print(f"✅ {code} (row {row}): {price}")
            cell = ws.cell(row=row, column=max_col, value=price)

            # ✅ 어제 데이터가 있다면 색상 비교
            if max_col >= 6:  # 종목코드, 종목명, + 2일치 이상
                try:
                    prev_value = ws.cell(row=row, column=max_col - 1).value
                    if prev_value is not None:
                        prev_value = float(prev_value)

                        # 🔒 기존 셀 색상이 없을 때만 적용
                        if not cell.font or not cell.font.color or cell.font.color.rgb is None:
                            if price > prev_value:
                                cell.font = Font(color="FF0000")  # 빨강
                            elif price < prev_value:
                                cell.font = Font(color="0000FF")  # 파랑
                except Exception:
                    pass
        else:
            print(f"⚠️ {code} (row {row}) not in closes")

    wb.save(file_path)
    print(f"✅ 기존 서식 유지하며 저장 완료: {file_path}")

def update_price_history_from_kiwoom(base_file_path: str):
    """코스피 100.xlsx 기준으로 종가 기록 누적 저장"""
    df_base = pd.read_excel(base_file_path, dtype={"종목코드": str})
    if "종목코드" not in df_base.columns:
        raise ValueError("❌ '종목코드' 컬럼이 필요합니다.")

    code_list = df_base["종목코드"].dropna().tolist()
    code_list = [f"{int(code):06d}" for code in code_list]  # 안전하게 6자리로

    token = get_token()
    price_map = get_today_closes(token, code_list)
    update_history_file(base_file_path, price_map)


if __name__ == "__main__":
    kosdaq_file_path = "data/tracking/코스닥_20250714_history.xlsx"
    kospi_file_path = "data/tracking/코스피_20250714_history.xlsx"

    update_price_history_from_kiwoom(kosdaq_file_path)
    update_price_history_from_kiwoom(kospi_file_path)

    message = f"📊 {TODAY} 종가 기록 완료!\n\n📎 첨부된 파일을 확인하세요."
    send_to_telegram(BOT_TOKEN, GROUP_ID, message)

    send_to_telegram(BOT_TOKEN , GROUP_ID ,message="코스닥 150 종가 확인" , file_path=kosdaq_file_path)
    send_to_telegram(BOT_TOKEN , GROUP_ID ,message="코스피 200 종가 확인" , file_path=kospi_file_path)